import streamlit as st
st.write(st.__version__)
import pandas as pd
import numpy as np
import jpholiday
import datetime
import calendar
import io
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

st.title("伊藤忠フォーマット変換アプリ")

# Google Drive から雛形ファイルをダウンロード
TEMPLATE_URL = "https://drive.google.com/uc?export=download&id=1qbej2PjuZavlUKxRdbQe1F1QoHnshkPb"

@st.cache_data(show_spinner=False)
def load_template_workbook():
    response = requests.get(TEMPLATE_URL)
    response.raise_for_status()
    in_memory_file = io.BytesIO(response.content)
    return load_workbook(in_memory_file)

# セッションステートでアップロードファイルを管理
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = []

uploaded = st.file_uploader(
    "ファイルをアップロード（複数可）",
    type=['xlsx', 'csv'],
    accept_multiple_files=True,
    key="file_uploader"
)

# 新規アップロードファイルをセッションステートに追加（重複なし）
if uploaded:
    for f in uploaded:
        exists = any((f.name == uf.name and f.size == uf.size) for uf in st.session_state.uploaded_files)
        if not exists:
            st.session_state.uploaded_files.append(f)

# アップロード済みファイル一覧表示
if st.session_state.uploaded_files:
    st.write("アップロード済みファイル:")
    for f in st.session_state.uploaded_files:
        st.write(f.name)

# クリアボタン
if st.button("アップロードファイルを一括クリア"):
    st.session_state["clear_files"] = True

if st.session_state.get("clear_files", False):
    st.session_state.pop("uploaded_files", None)
    st.session_state["clear_files"] = False
    try:
        st.experimental_rerun()
    except Exception as e:
        st.error(f"リロード処理でエラーが発生しました: {e}")


output_name = st.text_input("出力ファイル名（拡張子不要）", value="", help="例：cats_202406 ※必須")

run_button = st.button("✅ 実行")

def is_holiday(date):
    return date.weekday() >= 5 or jpholiday.is_holiday(date)

def init_monthly_data():
    return {
        'weekday': {month: [0]*48 for month in range(4, 16)},
        'holiday': {month: [0]*48 for month in range(4, 16)},
        'weekday_days': {month: 0 for month in range(4, 16)},
        'holiday_days': {month: 0 for month in range(4, 16)}
    }

def read_uploaded(file):
    if file.name.endswith('.csv'):
        df = pd.read_csv(file, header=4)
        df['Sheet'] = file.name
        return [df]
    else:
        xlsx = pd.ExcelFile(file)
        all_sheets = []
        for sheet_name in xlsx.sheet_names:
            df = pd.read_excel(xlsx, sheet_name=sheet_name, header=4)
            df['Sheet'] = sheet_name
            all_sheets.append(df)
        return all_sheets

if run_button:
    if not st.session_state.uploaded_files:
        st.warning("ファイルをアップロードしてください。")
        st.stop()

    if output_name.strip() == "":
        st.warning("出力ファイル名を入力してください。")
        st.stop()

    with st.spinner("Excel出力中(処理中)..."):
        monthly_data = init_monthly_data()
        latest_month_map = {}

        wb = load_template_workbook()
        ws_template: Worksheet = wb["コマ単位集計雛形（送電端）"]

        for file in st.session_state.uploaded_files:
            dataframes = read_uploaded(file)
            for df in dataframes:
                if df.empty:
                    continue

                df_columns = df.columns.tolist()
                dates = pd.to_datetime(df[df_columns[0]], errors='coerce')
                valid_dates = dates.dropna()

                if valid_dates.empty:
                    continue

                first_date = valid_dates.iloc[0]
                month_key = first_date.month if first_date.month >= 4 else first_date.month + 12

                if (month_key not in latest_month_map) or (first_date > latest_month_map[month_key]["date"]):
                    latest_month_map[month_key] = {
                        "df": df,
                        "file": file,
                        "sheet": df['Sheet'].iloc[0],
                        "date": first_date
                    }

        for m_key, info in latest_month_map.items():
            df = info["df"]
            file = info["file"]
            sheet_name = info["sheet"]

            df_columns = df.columns.tolist()
            used_dates = set()

            for _, row in df.iterrows():
                date = pd.to_datetime(row[df_columns[0]], errors='coerce')
                if pd.isnull(date):
                    continue

                mm = date.month
                month_index = mm if mm >= 4 else mm + 12
                key = 'holiday' if is_holiday(date) else 'weekday'

                date_str = date.strftime("%Y-%m-%d")
                if date_str not in used_dates:
                    monthly_data[key + '_days'][month_index] += 1
                    used_dates.add(date_str)

                for i in range(1, 49):
                    if i >= len(df_columns):
                        continue
                    colname = df_columns[i]
                    val = pd.to_numeric(row[colname], errors='coerce')
                    if not pd.isnull(val):
                        monthly_data[key][month_index][i - 1] += val

            df_with_header = pd.read_excel(file, sheet_name=sheet_name, header=None)
            output_sheet_name = info["date"].strftime("%Y%m")
            if output_sheet_name in wb.sheetnames:
                del wb[output_sheet_name]
            ws_data = wb.create_sheet(title=output_sheet_name)
            for r in df_with_header.itertuples(index=False):
                ws_data.append(r)

        for m in range(4, 16):
            col_idx = m - 1  # C=3（4月）
            col_letter = get_column_letter(col_idx)
            for i in range(48):
                ws_template[f"{col_letter}{4+i}"] = monthly_data['weekday'][m][i]
            ws_template[f"{col_letter}57"] = monthly_data['weekday_days'][m]

        for m in range(4, 16):
            col_idx = 17 + (m - 4)  # Q=17（4月）
            col_letter = get_column_letter(col_idx)
            for i in range(48):
                ws_template[f"{col_letter}{4+i}"] = monthly_data['holiday'][m][i]
            ws_template[f"{col_letter}57"] = monthly_data['holiday_days'][m]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    st.success("変換が完了しました！")
    st.download_button(
        label="📥 処理済みExcelをダウンロード",
        data=output,
        file_name=output_name.strip() + ".xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
